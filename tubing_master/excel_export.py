"""Excel (.xlsx) exports for BOM, quotation, summary, and process documents."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List

from tubing_master.project_history import TubingProjectRecord
from tubing_master.quotation import PROCESS_CHARGE_LABELS


def _ensure_xlsx(path: Path) -> Path:
    path = Path(path)
    if path.suffix.lower() != ".xlsx":
        path = path.with_suffix(".xlsx")
    return path


def _autosize_columns(ws, *, min_width: float = 9.0, max_width: float = 44.0) -> None:
    """Column widths tuned so tables fit A4 width when combined with fit-to-page."""
    from openpyxl.utils import get_column_letter

    for col_idx in range(1, (ws.max_column or 1) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, (ws.max_row or 1) + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            for line in str(val).splitlines():
                max_len = max(max_len, len(line))
        ws.column_dimensions[col_letter].width = min(max_width, max(min_width, max_len + 2))


def _configure_worksheet_a4_print(
    ws,
    *,
    landscape: bool | None = None,
    style_header: bool = True,
    set_print_area: bool = True,
) -> None:
    """Page setup, margins, fit-to-width, and light header styling for A4 printing."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.worksheet.properties import PageSetupProperties

    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    if landscape is None:
        landscape = max_col > 4

    _autosize_columns(ws)

    # Two-column summary sheets: give the value column more room.
    if max_col == 2:
        ws.column_dimensions["A"].width = min(36.0, ws.column_dimensions["A"].width or 18)
        ws.column_dimensions["B"].width = min(52.0, max(ws.column_dimensions["B"].width or 18, 28))

    ws.page_margins = PageMargins(
        left=0.5,
        right=0.5,
        top=0.75,
        bottom=0.75,
        header=0.3,
        footer=0.3,
    )
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = (
        ws.ORIENTATION_LANDSCAPE if landscape else ws.ORIENTATION_PORTRAIT
    )
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    ws.print_options.horizontalCentered = True

    if set_print_area:
        last_col = get_column_letter(max_col)
        ws.print_area = f"A1:{last_col}{max_row}"

    if style_header and max_row >= 1:
        first = ws.cell(row=1, column=1).value
        if first is not None and str(first).strip() != "":
            ws.print_title_rows = "1:1"
            header_fill = PatternFill("solid", fgColor="D9E1F2")
            bold = Font(bold=True)
            header_align = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            for cell in ws[1]:
                cell.font = bold
                cell.fill = header_fill
                cell.alignment = header_align

    for row in ws.iter_rows(min_row=2, max_row=max_row):
        for cell in row:
            if cell.value is None:
                continue
            text = str(cell.value)
            if len(text) > 36 or "\n" in text:
                prev = cell.alignment
                cell.alignment = Alignment(
                    horizontal=prev.horizontal if prev else "general",
                    vertical="top",
                    wrap_text=True,
                )


def _apply_a4_print_to_workbook(wb, *, skip_hidden: bool = True) -> None:
    for ws in wb.worksheets:
        if skip_hidden and getattr(ws, "sheet_state", None) == "hidden":
            continue
        _configure_worksheet_a4_print(ws)


def _quotation_stock_row(q: Dict[str, Any]) -> List[Any]:
    """First Quotation row: incoming material (qty × unit cost = materials line)."""
    materials_cost = float(q.get("materials_cost", 0.0) or 0.0)
    stock_mat = float(q.get("stock_material_cost", 0.0) or 0.0)
    ppk = float(q.get("price_per_kg", 0.0) or 0.0)
    mass_kg = stock_mat / ppk if ppk > 1e-12 else 0.0
    eff_uc = materials_cost / mass_kg if mass_kg > 1e-12 else materials_cost
    return ["Stock", "Incoming Stock Materials", "—", mass_kg, eff_uc, ""]


def _append_quotation_line_rows(ws, q: Dict[str, Any]) -> None:
    for ln in q.get("lines") or []:
        lk = str(ln.get("line_kind") or "pass")
        if lk == "pass" and ln.get("pass") is not None:
            ps = str(int(ln["pass"]))
        elif lk == "surcharge":
            ps = "—"
        elif lk == "process_charge":
            slot = str(ln.get("slot") or "")
            ps = PROCESS_CHARGE_LABELS.get(slot, slot or "—")
        else:
            p = ln.get("pass")
            ps = str(int(p)) if p is not None else ""
        desc = str(ln.get("description", ln.get("item", "")))
        tool = str(ln.get("dies", ""))
        qty = float(ln.get("qty", 1.0) or 1.0)
        uc = float(ln.get("unit_cost", 0.0) or 0.0)
        comm = str(ln.get("comments", ""))
        ws.append([ps, desc, tool, qty, uc, comm])


def _center_worksheet_cells(ws) -> None:
    """Center all sheet cells horizontally and vertically (wrap enabled)."""
    from openpyxl.styles import Alignment

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 1):
        for cell in row:
            cell.alignment = center


def _style_quotation_total_row(ws, row_idx: int) -> None:
    from openpyxl.styles import Font

    bold = Font(bold=True)
    for col in range(1, (ws.max_column or 1) + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = bold


def write_pass_bom_xlsx(path: Path, headers: List[str], rows: List[List[str]]) -> None:
    from openpyxl import Workbook

    path = _ensure_xlsx(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Pass BOM"
    ws.append(headers)
    for row in rows:
        ws.append(row)
    _configure_worksheet_a4_print(ws)
    _center_worksheet_cells(ws)
    wb.save(path)


def write_quotation_xlsx(path: Path, q: Dict[str, Any]) -> None:
    from openpyxl import Workbook

    path = _ensure_xlsx(path)
    wb = Workbook()

    # Reference sheet (visible); print the Quotation tab only (active sheet when the file opens).
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.append(["Field", "Value"])
    ws_sum.append(["Currency", str(q.get("currency", "USD"))])
    ws_sum.append(["Stock length (m)", float(q.get("stock_length_m", 0.0) or 0.0)])
    ws_sum.append(["Price per kg", float(q.get("price_per_kg", 0.0) or 0.0)])
    ws_sum.append(["Density (kg/m³)", float(q.get("density_kg_m3", 0.0) or 0.0)])
    ws_sum.append(
        [
            "Stock price (mass × price/kg)",
            float(q.get("stock_material_cost", 0.0) or 0.0),
        ]
    )
    ws_sum.append(["Additional cost", float(q.get("additional_cost", 0.0) or 0.0)])
    ws_sum.append(["Materials (stock + additional)", float(q.get("materials_cost", 0.0) or 0.0)])
    ws_sum.append(["Drawing charges", float(q.get("drawing_charges", 0.0) or 0.0)])
    ws_sum.append(["TOTAL", float(q.get("total", 0.0) or 0.0)])

    wl = wb.create_sheet("Quotation")
    wb.active = wl
    wl.append(["Pass", "Description", "Die", "Qty", "Unit cost", "Comments"])
    wl.append(_quotation_stock_row(q))
    _append_quotation_line_rows(wl, q)
    total_row = wl.max_row + 1
    wl.append(["Total", "", "—", "", float(q.get("total", 0.0) or 0.0), ""])
    _style_quotation_total_row(wl, total_row)

    _configure_worksheet_a4_print(ws_sum, set_print_area=False)
    _configure_worksheet_a4_print(wl)
    _center_worksheet_cells(wl)
    wb.save(path)


def write_process_document_xlsx(path: Path, rec: TubingProjectRecord, history_file: Path | str) -> None:
    """Pass schedule / Fetch History Excel export: Overview + Pass BOM sheets only."""
    from openpyxl import Workbook

    path = _ensure_xlsx(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Overview"
    ws.append(["Field", "Value"])
    for row in [
        ("Title", rec.title),
        ("Record id", rec.id),
        ("Saved", rec.saved_at),
        ("History / reference file", str(history_file)),
        ("Incoming OD (mm)", rec.in_od_mm),
        ("Incoming ID (mm)", rec.in_id_mm),
        ("Target OD (mm)", rec.out_od_mm),
        ("Target ID (mm)", rec.out_id_mm),
        ("Material", rec.material),
        ("Drawing method", rec.drawing_method),
    ]:
        ws.append(list(row))

    wbom = wb.create_sheet("Pass BOM")
    if rec.pass_bom:
        bom = rec.pass_bom
        detail = bom.get("detail_rows")
        alt_lines = bom.get("lines")
        tab = detail if isinstance(detail, list) and detail else alt_lines
        if isinstance(tab, list) and tab and all(isinstance(r, dict) for r in tab):
            keys = list(tab[0].keys())
            wbom.append(keys)
            for r in tab:
                wbom.append([r.get(k, "") for k in keys])
        else:
            wbom.append(["Full JSON"])
            wbom.append([json.dumps(bom, indent=2)])
    else:
        wbom.append(["(no BOM snapshot)"])

    _apply_a4_print_to_workbook(wb)
    _center_worksheet_cells(wbom)
    wb.save(path)
